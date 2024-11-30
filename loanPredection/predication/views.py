import pandas as pd
from django.http import JsonResponse
from .models import Loan, Customer
from django.views.decorators.csrf import csrf_exempt
import openpyxl
import json
from datetime import date, timedelta   
import random
from io import BytesIO


@csrf_exempt
def upload_loan_data(request):
    if request.method == "POST" and request.FILES.get("file"):
        file = request.FILES["file"]
        try:
            # Load the Excel file
            df = pd.read_excel(file)

            # Iterate over rows and save each record to the database
            for _, row in df.iterrows():
                _, created = Loan.objects.get_or_create(
                    loan_id=row["Loan ID"],  # Use loan_id as the unique identifier
                    defaults={
                        "customer_id": row["Customer ID"],
                        "loan_amount": row["Loan Amount"],
                        "tenure": row["Tenure"],
                        "interest_rate": row["Interest Rate"],
                        "monthly_payment": row["Monthly payment"],
                        "emis_paid_on_time": row["EMIs paid on Time"],
                        "date_of_approval": row["Date of Approval"],
                        "end_date": row["End Date"],
                    },
                )
            return JsonResponse({"message": "Data uploaded successfully"}, status=201)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
    return JsonResponse({"error": "Invalid request"}, status=400)


@csrf_exempt
def upload_customer_data(request):
    if request.method == "POST":
        try:
            # Check if file exists in request
            if 'excel_file' not in request.FILES:
                return JsonResponse({
                    'error': 'No file uploaded. Please upload a file with key "excel_file"'
                }, status=400)

            excel_file = request.FILES['excel_file']
            
            # Validate file extension
            if not excel_file.name.endswith(('.xls', '.xlsx')):
                return JsonResponse({
                    'error': 'Invalid file format. Please upload an Excel file (.xls or .xlsx)'
                }, status=400)

            # Load the Excel file
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            # Get headers to verify column names
            headers = [cell.value for cell in sheet[1]]
            required_columns = ['Customer ID', 'First Name', 'Last Name', 'Age', 
                              'Phone Number', 'Monthly Salary', 'Approved Limit']
            
            # Verify all required columns are present
            missing_columns = [col for col in required_columns if col not in headers]
            if missing_columns:
                return JsonResponse({
                    'error': f'Missing required columns: {", ".join(missing_columns)}'
                }, status=400)

            # Track successful and failed records
            success_count = 0
            failed_records = []

            # Skip the header row and iterate through the data
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                try:
                    customer_id, first_name, last_name, age, phone_number, monthly_salary, approved_limit = row
                    
                    # Basic validation
                    if not all([first_name, last_name, age, phone_number, monthly_salary]):
                        failed_records.append(f"Row {row_num}: Missing required fields")
                        continue

                    # Create or update customer
                    Customer.objects.update_or_create(
                        phone_number=phone_number,  # Using phone_number as unique identifier
                        defaults={
                            'first_name': first_name,
                            'last_name': last_name,
                            'age': age,
                            'monthly_salary': monthly_salary,
                            'approved_limit': approved_limit if approved_limit else round((36 * float(monthly_salary)) / 100000) * 100000,
                        }
                    )
                    success_count += 1

                except Exception as e:
                    failed_records.append(f"Row {row_num}: {str(e)}")

            # Prepare response
            response_data = {
                'message': f'Successfully processed {success_count} records',
                'total_rows': sheet.max_row - 1,  # Excluding header
                'successful_records': success_count,
                'failed_records': len(failed_records),
            }
            
            if failed_records:
                response_data['errors'] = failed_records

            status_code = 200 if success_count > 0 else 400
            return JsonResponse(response_data, status=status_code)

        except Exception as e:
            return JsonResponse({
                'error': f'Error processing file: {str(e)}'
            }, status=400)

    return JsonResponse({'error': 'Invalid request method'}, status=405)


# function to register a new customer 
@csrf_exempt
def add_customer(request):  
    if request.method == "POST":
        try:
            # Parse JSON request body
            data = json.loads(request.body)
            print("Received data:", data)

            # Extract fields from request body
            first_name = data.get("first_name")
            last_name = data.get("last_name")
            age = data.get("age")
            monthly_income = data.get("monthly_income")
            phone_number = data.get("phone_number")

            # Validate required fields
            if not all([first_name, last_name, age, monthly_income, phone_number]):
                missing_fields = [field for field in ["first_name", "last_name", "age", "monthly_income", "phone_number"] 
                                if not data.get(field)]
                return JsonResponse({
                    "error": "Missing required fields",
                    "missing_fields": missing_fields
                }, status=400)

            # Check if phone number already exists
            if Customer.objects.filter(phone_number=phone_number).exists():
                return JsonResponse({"error": "Phone number already registered"}, status=409)

            # Calculate approved_limit (rounded to nearest lakh)
            approved_limit = round((36 * float(monthly_income)) / 100000) * 100000

            # Create new customer in the database
            customer = Customer.objects.create(
                first_name=first_name,
                last_name=last_name,
                age=age,
                monthly_salary=monthly_income,
                phone_number=phone_number,
                approved_limit=approved_limit
            )

            # Return success response with required fields
            return JsonResponse({
                "customer_id": customer.customer_id,
                "name": f"{customer.first_name} {customer.last_name}",
                "age": customer.age,
                "monthly_income": customer.monthly_salary,
                "approved_limit": customer.approved_limit,
                "phone_number": customer.phone_number
            }, status=201)

        except json.JSONDecodeError as e:
            print("JSON Decode Error:", str(e))
            return JsonResponse({"error": "Invalid JSON data"}, status=400)
        except Exception as e:
            print("Exception:", str(e))
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Method not allowed"}, status=405)


# Function to check the eligibility criteria od the customer
@csrf_exempt
def loan_eligibility(request):
    if request.method == "POST":
        try:
            # Parse JSON request body
            data = json.loads(request.body)
            print("Received data:", data)  # Debug print
            
            # Extract data from the request with proper error handling
            customer_id = data.get('customer_id')
            loan_amount = data.get('loan_amount')
            interest_rate = data.get('interest_rate')
            tenure = data.get('tenure')

            # Validate all required fields are present
            if None in [customer_id, loan_amount, interest_rate, tenure]:
                return JsonResponse({
                    "error": "Missing required fields",
                    "required_fields": {
                        "customer_id": customer_id is not None,
                        "loan_amount": loan_amount is not None,
                        "interest_rate": interest_rate is not None,
                        "tenure": tenure is not None
                    }
                }, status=400)

            # Validate data types
            try:
                customer_id = int(customer_id)
                loan_amount = float(loan_amount)
                interest_rate = float(interest_rate)
                tenure = int(tenure)
            except (ValueError, TypeError):
                return JsonResponse({
                    "error": "Invalid data types. Please ensure:\n" +
                            "customer_id: integer\n" +
                            "loan_amount: number\n" +
                            "interest_rate: number\n" +
                            "tenure: integer"
                }, status=400)

            # Fetch customer details
            try:
                customer = Customer.objects.get(customer_id=customer_id)
            except Customer.DoesNotExist:
                return JsonResponse({
                    "error": f"Customer with ID {customer_id} not found"
                }, status=404)

            # Get all loans for this customer
            customer_loans = Loan.objects.filter(customer_id=customer_id)
            
            # Calculate credit score components
            past_loans_paid_on_time = sum(loan.emis_paid_on_time for loan in customer_loans)
            no_of_loans_taken = customer_loans.count()
            current_year = date.today().year
            loan_activity_current_year = customer_loans.filter(
                date_of_approval__year=current_year
            ).count()
            loan_approved_volume = sum(float(loan.loan_amount) for loan in customer_loans)

            # Calculate total current loans
            total_current_loans = sum(float(loan.loan_amount) for loan in customer_loans)

            # Credit score calculation
            if total_current_loans > float(customer.approved_limit):
                credit_score = 0
            else:
                credit_score = (
                    past_loans_paid_on_time * 0.4 +
                    no_of_loans_taken * 0.3 +
                    loan_activity_current_year * 0.2 +
                    loan_approved_volume * 0.1
                )

            # Determine approval status and adjusted interest rate
            corrected_interest_rate = interest_rate
            approved = False

            if credit_score > 50:
                approved = True
            elif 30 < credit_score <= 50:
                approved = True
                corrected_interest_rate = max(interest_rate, 12)
            elif 10 < credit_score <= 30:
                approved = True
                corrected_interest_rate = max(interest_rate, 16)
            else:
                approved = False
                corrected_interest_rate = max(interest_rate, 16)

            # Calculate monthly installment (EMI)
            monthly_rate = corrected_interest_rate / 100 / 12
            monthly_installment = (
                loan_amount * monthly_rate / (1 - (1 + monthly_rate) ** (-tenure * 12))
            )

            # Check monthly EMI constraint
            if monthly_installment > 0.5 * float(customer.monthly_salary):
                approved = False

            # Construct the response
            response = {
                "customer_id": customer_id,
                "approval": approved,
                "interest_rate": interest_rate,
                "corrected_interest_rate": corrected_interest_rate,
                "tenure": tenure,
                "monthly_installment": round(monthly_installment, 2) if approved else 0.0,
            }
            return JsonResponse(response, status=200)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON data"}, status=400)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Method not allowed"}, status=405)



# creating a new loan against a a customer

@csrf_exempt
def create_new_loan(request):
    if request.method == "POST":
        try:
            # Parse JSON request body
            data = json.loads(request.body)
            print("Received data:", data)  # Debug print

            # Extract fields from request
            customer_id = data.get('customer_id')
            loan_amount = data.get('loan_amount')
            interest_rate = data.get('interest_rate')
            tenure = data.get('tenure')

            # Validate required fields
            if None in [customer_id, loan_amount, interest_rate, tenure]:
                return JsonResponse({
                    "error": "Missing required fields",
                    "required_fields": {
                        "customer_id": customer_id is not None,
                        "loan_amount": loan_amount is not None,
                        "interest_rate": interest_rate is not None,
                        "tenure": tenure is not None
                    }
                }, status=400)

            # Check customer exists
            try:
                customer = Customer.objects.get(customer_id=customer_id)
            except Customer.DoesNotExist:
                return JsonResponse({
                    "error": f"Customer with ID {customer_id} not found"
                }, status=404)

            # Prepare data for eligibility check
            eligibility_data = {
                "customer_id": customer_id,
                "loan_amount": loan_amount,
                "interest_rate": interest_rate,
                "tenure": tenure
            }

            # Create a mock request with proper wsgi.input
            class MockRequest:
                def __init__(self, data):
                    self.method = 'POST'
                    self._body = json.dumps(data).encode('utf-8')
                    self.META = {
                        'wsgi.input': BytesIO(self._body),
                        'wsgi.url_scheme': 'http',
                        'SERVER_NAME': 'localhost',
                        'SERVER_PORT': '8000',
                    }

                @property
                def body(self):
                    return self._body

            # Check eligibility
            mock_request = MockRequest(eligibility_data)
            eligibility_response = loan_eligibility(mock_request)
            eligibility_result = json.loads(eligibility_response.content)

            print("Eligibility result:", eligibility_result)  # Debug print

            # Check if loan is approved
            if not eligibility_result.get("approval"):
                return JsonResponse({
                    "loan_id": None,
                    "customer_id": customer_id,
                    "loan_approved": False,
                    "message": "Loan not approved due to eligibility criteria",
                    "monthly_installment": 0.0
                }, status=200)

            # If approved, create new loan
            try:
                # Generate unique loan ID
                while True:
                    loan_id = random.randint(10000, 99999)
                    if not Loan.objects.filter(loan_id=loan_id).exists():
                        break

                # Create and save new loan
                new_loan = Loan.objects.create(
                    customer_id=customer_id,
                    loan_id=loan_id,
                    loan_amount=loan_amount,
                    tenure=tenure,
                    interest_rate=eligibility_result.get("corrected_interest_rate", interest_rate),
                    monthly_payment=eligibility_result.get("monthly_installment", 0),
                    emis_paid_on_time=0,
                    date_of_approval=date.today(),
                    end_date=date.today() + timedelta(days=tenure * 365)
                )

                return JsonResponse({
                    "loan_id": new_loan.loan_id,
                    "customer_id": customer_id,
                    "loan_approved": True,
                    "message": "Loan approved and created successfully",
                    "monthly_installment": eligibility_result.get("monthly_installment", 0)
                }, status=201)

            except Exception as e:
                print(f"Error creating loan: {str(e)}")  # Debug print
                return JsonResponse({
                    "error": f"Error creating loan: {str(e)}"
                }, status=500)

        except json.JSONDecodeError:
            return JsonResponse({
                "error": "Invalid JSON data"
            }, status=400)
        except Exception as e:
            print(f"General error: {str(e)}")  # Debug print
            return JsonResponse({
                "error": str(e)
            }, status=500)

    return JsonResponse({
        "error": "Method not allowed"
    }, status=405)



# view loan details against a particular loan-id

@csrf_exempt
def view_loan_against_loan_id(request, loan_id):
    print(f"Received request for loan ID: {loan_id}")  # Debug print
    if request.method == "GET":
        try:
            loan = Loan.objects.get(loan_id=loan_id)
            # Construct the response
            response = {
                "loan_id": loan.loan_id,
                "customer_id": loan.customer_id,
                "loan_amount": float(loan.loan_amount),
                "interest_rate": loan.interest_rate,
                "monthly_payment": float(loan.monthly_payment),
                "tenure": loan.tenure,
                "date_of_approval": loan.date_of_approval,
                "end_date": loan.end_date,
            }
            return JsonResponse(response, status=200)

        except Loan.DoesNotExist:
            print(f"Loan with ID {loan_id} not found")  # Debug print
            return JsonResponse({"error": f"Loan with ID {loan_id} not found"}, status=404)
        except Exception as e:
            print(f"Error: {str(e)}")  # Debug print
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Method not allowed"}, status=405)




# function for getting loan details against a customer id
@csrf_exempt
def view_loan_against_customer_id(request, customer_id):
    if request.method == "GET":
        try:
            # Fetch all loans for the given customer_id
            loans = Loan.objects.filter(customer_id=customer_id)

            if not loans.exists():
                return JsonResponse({
                    "error": f"No loans found for customer ID {customer_id}"
                }, status=404)

            # Prepare the response data
            loan_items = []
            for loan in loans:
                # Calculate repayments left
                today = date.today()
                months_since_approval = (today.year - loan.date_of_approval.year) * 12 + (today.month - loan.date_of_approval.month)
                repayments_left = max(loan.tenure * 12 - months_since_approval, 0)

                loan_items.append({
                    "loan_id": loan.loan_id,
                    "loan_amount": float(loan.loan_amount),
                    "interest_rate": loan.interest_rate,
                    "monthly_installment": float(loan.monthly_payment),
                    "repayments_left": repayments_left
                })

            return JsonResponse(loan_items, safe=False, status=200)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Method not allowed"}, status=405)
